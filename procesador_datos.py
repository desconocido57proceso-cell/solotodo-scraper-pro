import pandas as pd
import re
from openpyxl.utils import get_column_letter 

def limpiar_y_clasificar(df_sucio, nombre_final_excel):
    print(f"🧹 Refinería M4: Aplicando estética y filtros...")
    
    # 1. Filtro de pureza Lenovo
    df = df_sucio[df_sucio['Modelo'].str.contains('Lenovo', case=False, na=False)].copy()

    # 2. Extracción con Regex (Tu especialidad)
    def extraer_ram(texto):
        match = re.search(r'(\d+)\s*GB', str(texto))
        return match.group(0) if match else "Verificar en Web"

    def extraer_procesador(texto):
        patron = r'(i\d|Ryzen \d|M\d|Apple Silicon)'
        match = re.search(patron, str(texto), re.IGNORECASE)
        return match.group(0) if match else "Verificar Specs"

    df['RAM'] = df['Modelo'].apply(extraer_ram)
    df['Procesador'] = df['Modelo'].apply(extraer_procesador)
    df = df.sort_values(by='Precio_CLP', ascending=True)

    # 3. --- INGENIERÍA DE DISEÑO EXCEL (ACTUALIZADO PARA USAR LA FUNCIÓN) ---
    with pd.ExcelWriter(nombre_final_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte_Lenovo')
        worksheet = writer.sheets['Reporte_Lenovo']
        
        # EL TRUCO SENIOR: Usamos enumerate para saber el número de columna (1, 2, 3...)
        # y la función get_column_letter la convierte en letra (A, B, C...)
        for i, col in enumerate(worksheet.columns, 1): 
            max_length = 0
            
            # ACTIVACIÓN: Aquí es donde usamos la función que estaba en gris
            column_letter = get_column_letter(i) 
            
            for cell in col:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except: 
                    pass
            
            # Ajustamos el ancho para que no se vea colapsado
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"✅ BINGO! Reporte Ejecutivo y ESPACIOSO listo: {nombre_final_excel}")